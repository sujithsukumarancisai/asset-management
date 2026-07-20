import io
from datetime import date, datetime
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

from sqlalchemy.orm import Session

from app.models.employee import Employee
from app.models.asset import Asset
from app.models.assignment import Assignment

HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _style_header(ws, columns: List[str]):
    ws.append(columns)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for i, col in enumerate(columns, start=1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A"].width = max(16, len(col) + 4)


def _build_rows(report_type: str, db: Session):
    if report_type == "employee":
        columns = ["Emp Code", "Name", "Department", "Designation", "Status", "Assets Assigned"]
        rows = []
        for e in db.query(Employee).order_by(Employee.emp_code).all():
            count = sum(1 for a in e.assignments if a.status == "Active")
            rows.append([e.emp_code, e.name, e.department, e.designation or "-", e.status, count])
        return columns, rows

    if report_type == "asset":
        columns = ["Asset Code", "Name", "Category", "Brand", "Model", "Serial No", "Status", "Condition"]
        rows = []
        for a in db.query(Asset).order_by(Asset.asset_code).all():
            rows.append([a.asset_code, a.name, a.category, a.brand or "-", a.model or "-",
                         a.serial_number or "-", a.status, a.condition])
        return columns, rows

    if report_type == "department":
        columns = ["Department", "Employee Count", "Assets Assigned"]
        depts = {}
        for e in db.query(Employee).all():
            d = depts.setdefault(e.department, {"employees": 0, "assets": 0})
            d["employees"] += 1
            d["assets"] += sum(1 for a in e.assignments if a.status == "Active")
        rows = [[name, v["employees"], v["assets"]] for name, v in sorted(depts.items())]
        return columns, rows

    if report_type == "available":
        columns = ["Asset Code", "Name", "Category", "Brand", "Condition"]
        rows = [
            [a.asset_code, a.name, a.category, a.brand or "-", a.condition]
            for a in db.query(Asset).filter(Asset.status == "Available").order_by(Asset.asset_code).all()
        ]
        return columns, rows

    if report_type == "assigned":
        columns = ["Asset Code", "Asset Name", "Employee", "Department", "Assigned Date"]
        rows = []
        for a in db.query(Assignment).filter(Assignment.status == "Active").all():
            rows.append([
                a.asset.asset_code, a.asset.name, a.employee.name, a.employee.department,
                a.assigned_date.strftime("%Y-%m-%d") if a.assigned_date else "-",
            ])
        return columns, rows

    if report_type == "maintenance":
        columns = ["Asset Code", "Name", "Category", "Brand", "Condition"]
        rows = [
            [a.asset_code, a.name, a.category, a.brand or "-", a.condition]
            for a in db.query(Asset).filter(Asset.status == "Maintenance").order_by(Asset.asset_code).all()
        ]
        return columns, rows

    if report_type == "warranty":
        columns = ["Asset Code", "Name", "Brand", "Warranty Expiry", "Status"]
        rows = []
        today = date.today()
        for a in db.query(Asset).filter(Asset.warranty_expiry.isnot(None)).order_by(Asset.warranty_expiry).all():
            status = "Expired" if a.warranty_expiry < today else "Valid"
            rows.append([a.asset_code, a.name, a.brand or "-", a.warranty_expiry.strftime("%Y-%m-%d"), status])
        return columns, rows

    raise ValueError(f"Unknown report type: {report_type}")


REPORT_TITLES = {
    "employee": "Employee Report",
    "asset": "Asset Report",
    "department": "Department Report",
    "available": "Available Assets Report",
    "assigned": "Assigned Assets Report",
    "maintenance": "Maintenance Report",
    "warranty": "Warranty Expiry Report",
}


def generate_excel(report_type: str, db: Session) -> io.BytesIO:
    columns, rows = _build_rows(report_type, db)
    wb = Workbook()
    ws = wb.active
    ws.title = REPORT_TITLES.get(report_type, "Report")[:31]
    _style_header(ws, columns)
    for row in rows:
        ws.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_pdf(report_type: str, db: Session) -> io.BytesIO:
    columns, rows = _build_rows(report_type, db)
    title = REPORT_TITLES.get(report_type, "Report")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        topMargin=1.5 * cm, bottomMargin=1.5 * cm, leftMargin=1.5 * cm, rightMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(f"Company Asset Management — {title}", styles["Title"]),
        Paragraph(f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles["Normal"]),
        Spacer(1, 0.5 * cm),
    ]

    data = [columns] + [[str(c) for c in row] for row in rows]
    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E293B")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F5F9")]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    elements.append(table)
    doc.build(elements)
    buf.seek(0)
    return buf
